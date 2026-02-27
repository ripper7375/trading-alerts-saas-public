"""
import_to_excel.py
-----------------
Reads every .txt (tab-separated) file in this folder and writes each one
into a separate sheet of a single Excel workbook saved in the same folder.

Sheet order and names are defined in SHEET_MAP below.
Filename matching uses startswith() so timestamped filenames keep working
even after the timestamp portion changes.

Requirements:
    pip install openpyxl

Output:
    XAUUSD_M5_AllData.xlsx  (created in the same folder as this script)
"""

import csv
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit(
        "openpyxl is not installed.\n"
        "Run:  pip install openpyxl"
    )

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).parent
OUTPUT_FILE = BASE_DIR / "XAUUSD_M5_AllData.xlsx"

# (filename_prefix, sheet_name) — order here = sheet tab order in Excel
SHEET_MAP = [
    ("Dual_TEMA_HL",         "Dual_Tema"),
    ("FractalDiagonals",     "Diagonals"),
    ("FractalTrendlines",    "Trendlines"),
    ("KC_ATF",               "KC"),
    ("PinbarDetector",       "Pinbar"),
    ("PriceData",            "Price"),
    ("SR_Levels",            "SR_Levels"),
    ("TEMA_HRMA_SMMA",       "Tema_Hrma_Smma"),
    ("XAUUSD_M5_HA_Export",  "HA"),
    ("ZigZag",               "ZigZag"),
    ("ZScore",               "ZScore"),
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def find_txt_for_prefix(prefix: str, txt_files: list[Path]) -> Path | None:
    """Return the first .txt file whose stem starts with the given prefix."""
    for path in txt_files:
        if path.stem.startswith(prefix):
            return path
    return None


def import_txt_to_excel() -> None:
    txt_files = sorted(BASE_DIR.glob("*.txt"))
    if not txt_files:
        print("No .txt files found in", BASE_DIR)
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    for prefix, sheet_name in SHEET_MAP:
        txt_path = find_txt_for_prefix(prefix, txt_files)

        if txt_path is None:
            print(f"  [SKIP]  No file found matching prefix '{prefix}'")
            continue

        ws = wb.create_sheet(title=sheet_name)

        with open(txt_path, encoding="utf-8", newline="") as fh:
            reader = csv.reader(fh, delimiter="\t")
            for row in reader:
                ws.append(row)

        print(f"  [OK]    {txt_path.name}  ->  sheet '{sheet_name}'  ({ws.max_row} rows, {ws.max_column} cols)")

    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print(f"Importing .txt files from: {BASE_DIR}\n")
    import_txt_to_excel()
