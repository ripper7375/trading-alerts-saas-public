import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

v6_path = r"docs\migration-orders\prompt-to-claude-code\antigravity\migration-process-handbook-antigravity-v6.xlsx"
v7_path = r"docs\migration-orders\prompt-to-claude-code\antigravity\migration-process-handbook-antigravity-v7.xlsx"

wb = openpyxl.load_workbook(v6_path)

# Helper function to duplicate a row and insert it right after
def insert_and_copy_row(ws, source_row_idx, new_row_idx):
    ws.insert_rows(new_row_idx)
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=source_row_idx, column=col)
        dst_cell = ws.cell(row=new_row_idx, column=col)
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell.font = src_cell.font.copy()
            dst_cell.border = src_cell.border.copy()
            dst_cell.fill = src_cell.fill.copy()
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = src_cell.protection.copy()
            dst_cell.alignment = src_cell.alignment.copy()

print("1. Updating Audit_Report sheet...")
ws_audit = wb["Audit_Report"]
for r in range(1, ws_audit.max_row + 1):
    for c in range(1, ws_audit.max_column + 1):
        val = ws_audit.cell(r, c).value
        if val and isinstance(val, str):
            if "4A-W3," in val:
                ws_audit.cell(r, c).value = val.replace("4A-W3,", "4A-W3a, 4A-W3b,")
            elif "4A-W3/" in val:
                ws_audit.cell(r, c).value = val.replace("4A-W3/", "4A-W3a/W3b/")

print("2. Updating Runway sheet...")
ws_runway = wb["Runway"]
row_w3 = None
for r in range(1, ws_runway.max_row + 1):
    if ws_runway.cell(r, 1).value == "4A-W3":
        row_w3 = r
        break

if row_w3:
    ws_runway.cell(row_w3, 1).value = "4A-W3a"
    if ws_runway.max_column >= 2 and ws_runway.cell(row_w3, 2).value:
        ws_runway.cell(row_w3, 2).value = "BUILD Wise recipient onboarding backend (money-service)"
    insert_and_copy_row(ws_runway, row_w3, row_w3 + 1)
    ws_runway.cell(row_w3 + 1, 1).value = "4A-W3b"
    if ws_runway.max_column >= 2 and ws_runway.cell(row_w3 + 1, 2).value:
        ws_runway.cell(row_w3 + 1, 2).value = "BUILD Wise recipient onboarding UI (monolith)"

print("3. Updating Rev2_Changes sheet...")
ws_rev2 = wb["Rev2_Changes"]
row_rev2_w3 = None
for r in range(1, ws_rev2.max_row + 1):
    if ws_rev2.cell(r, 1).value == "4A-W3":
        row_rev2_w3 = r
        break

if row_rev2_w3:
    ws_rev2.cell(row_rev2_w3, 1).value = "4A-W3a"
    if ws_rev2.max_column >= 2 and ws_rev2.cell(row_rev2_w3, 2).value:
        ws_rev2.cell(row_rev2_w3, 2).value = "4A-W3a"
    insert_and_copy_row(ws_rev2, row_rev2_w3, row_rev2_w3 + 1)
    ws_rev2.cell(row_rev2_w3 + 1, 1).value = "4A-W3b"
    if ws_rev2.max_column >= 2 and ws_rev2.cell(row_rev2_w3 + 1, 2).value:
        ws_rev2.cell(row_rev2_w3 + 1, 2).value = "4A-W3b"

print("4. Updating Roles sheet...")
ws_roles = wb["Roles"]
for r in range(1, ws_roles.max_row + 1):
    val = ws_roles.cell(r, 1).value
    if val and isinstance(val, str) and "4A-W3/" in val:
        ws_roles.cell(r, 1).value = val.replace("4A-W3/", "4A-W3a/W3b/")

print("5. Updating Guide sheet...")
ws_guide = wb["Guide"]
for r in range(1, ws_guide.max_row + 1):
    for c in range(1, ws_guide.max_column + 1):
        val = ws_guide.cell(r, c).value
        if val and isinstance(val, str) and "4A-W3," in val:
            ws_guide.cell(r, c).value = val.replace("4A-W3,", "4A-W3a, 4A-W3b,")

print("6. Updating Task_Description sheet...")
ws_task = wb["Task_Description"]
row_task_w3 = None
for r in range(1, ws_task.max_row + 1):
    if ws_task.cell(r, 2).value == "4A-W3":
        row_task_w3 = r
        break

if row_task_w3:
    ws_task.cell(row_task_w3, 2).value = "4A-W3a"
    if ws_task.max_column >= 6 and ws_task.cell(row_task_w3, 6).value:
        ws_task.cell(row_task_w3, 6).value = "Part 19.5: BUILD Wise recipient onboarding backend (money-service)"
    insert_and_copy_row(ws_task, row_task_w3, row_task_w3 + 1)
    ws_task.cell(row_task_w3 + 1, 2).value = "4A-W3b"
    if ws_task.max_column >= 6 and ws_task.cell(row_task_w3 + 1, 6).value:
        ws_task.cell(row_task_w3 + 1, 6).value = "Part 19.5: BUILD Wise recipient onboarding UI (monolith)"

print("7. Updating Instruction sheet...")
ws_inst = wb["Instruction"]
row_inst_w3 = None
for r in range(1, ws_inst.max_row + 1):
    if ws_inst.cell(r, 2).value == "4A-W3":
        row_inst_w3 = r
        break

if row_inst_w3:
    # Update 4A-W3 to 4A-W3a
    ws_inst.cell(row_inst_w3, 2).value = "4A-W3a"
    ws_inst.cell(row_inst_w3, 6).value = "Part 19.5: BUILD Wise recipient onboarding backend (money-service)"
    ws_inst.cell(row_inst_w3, 7).value = (
        "Here's the PRE-DRAFT from session 4A-W2 at docs/migration-orders/4a-w3a-wise-recipient-backend.migration-order.md "
        "— produce the DRAFT for session 4A-W3a per 00-SKELETON-AND-RULES.md. Variant: PORT rules (dial LOW). "
        "Seed from 04-...§4 (4A-W3a) and 02-wise-platform-api-integration-reference.md §4.2-4.3. "
        "Invariants: schema-driven requirements endpoint, PII redaction (accountTail & detailsFingerprint SHA-256 only), native fetch and crypto only."
    )
    ws_inst.cell(row_inst_w3, 11).value = (
        "Read CLAUDE.md and docs/migration-orders/EXECUTOR-PROTOCOL.md. CONFIRM the APPROVED order for session 4A-W3a against "
        "the current codebase AND runtime state, and show me: what changed since drafting, the \"done when\" checks, and any failing entry criterion. "
        "Do not execute until I say go."
    )
    ws_inst.cell(row_inst_w3, 17).value = (
        "Wrap up per EXECUTOR-PROTOCOL §3: tests + results, fill Deviations, update the artifacts, harvest any lesson into LESSONS-LEARNED.md, "
        "then PRE-DRAFT session 4A-W3b's order and show it to me."
    )

    # Insert 4A-W3b
    insert_and_copy_row(ws_inst, row_inst_w3, row_inst_w3 + 1)
    row_w3b = row_inst_w3 + 1
    ws_inst.cell(row_w3b, 2).value = "4A-W3b"
    ws_inst.cell(row_w3b, 6).value = "Part 19.5: BUILD Wise recipient onboarding UI (monolith)"
    ws_inst.cell(row_w3b, 7).value = (
        "Here's the PRE-DRAFT from session 4A-W3a at docs/migration-orders/4a-w3b-wise-recipient-ui.migration-order.md "
        "— produce the DRAFT for session 4A-W3b per 00-SKELETON-AND-RULES.md. Variant: UI-BUILD rules (dial HIGH). "
        "Seed from 04-...§4 (4A-W3b) and 4A-W3a backend endpoints."
    )
    ws_inst.cell(row_w3b, 11).value = (
        "Read CLAUDE.md and docs/migration-orders/EXECUTOR-PROTOCOL.md. CONFIRM the APPROVED order for session 4A-W3b against "
        "the current codebase AND runtime state, and show me: what changed since drafting, the \"done when\" checks, and any failing entry criterion. "
        "Do not execute until I say go."
    )
    ws_inst.cell(row_w3b, 17).value = (
        "Wrap up per EXECUTOR-PROTOCOL §3: tests + results, fill Deviations, update the artifacts, harvest any lesson into LESSONS-LEARNED.md, "
        "then PRE-DRAFT session 4A-W4's order and show it to me."
    )

    # Update 4A-W4 row's Col 7 handoff reference if 4A-W4 comes right after
    for r in range(row_w3b + 1, ws_inst.max_row + 1):
        if ws_inst.cell(r, 2).value == "4A-W4":
            col7_val = str(ws_inst.cell(r, 7).value or "")
            if "session 4A-W3" in col7_val:
                ws_inst.cell(r, 7).value = col7_val.replace("session 4A-W3", "session 4A-W3b")
            break

print("8. Updating Chat_Grouping sheet...")
ws_chat = wb["Chat_Grouping"]
row_chat_w3 = None
for r in range(1, ws_chat.max_row + 1):
    if ws_chat.cell(r, 2).value == "4A-W3":
        row_chat_w3 = r
        break

if row_chat_w3:
    ws_chat.cell(row_chat_w3, 2).value = "4A-W3a"
    insert_and_copy_row(ws_chat, row_chat_w3, row_chat_w3 + 1)
    ws_chat.cell(row_chat_w3 + 1, 2).value = "4A-W3b"

for r in range(1, ws_chat.max_row + 1):
    for c in range(1, ws_chat.max_column + 1):
        val = ws_chat.cell(r, c).value
        if val and isinstance(val, str) and "4A-W3," in val:
            ws_chat.cell(r, c).value = val.replace("4A-W3,", "4A-W3a, 4A-W3b,")

wb.save(v7_path)
print(f"Successfully saved updated Excel file to: {v7_path}")
