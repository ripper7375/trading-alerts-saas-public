import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

v6_path = r"docs\migration-orders\prompt-to-claude-code\antigravity\migration-process-handbook-antigravity-v6.xlsx"
wb = openpyxl.load_workbook(v6_path)

for sname in ['Task_Description', 'Instruction']:
    ws = wb[sname]
    print(f"\n=== Sheet: {sname} ===")
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v and isinstance(v, str) and '4A-W' in v for v in row_vals):
            print(f"\nRow {r}:")
            for idx, val in enumerate(row_vals):
                if val:
                    print(f"  Col {idx+1}: {val}")
