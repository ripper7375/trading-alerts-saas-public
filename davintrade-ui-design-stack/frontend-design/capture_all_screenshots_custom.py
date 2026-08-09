import asyncio
import os
import re
import sys
import time
import urllib.parse
import openpyxl
from playwright.async_api import async_playwright

EXCEL_PATH = r"D:\SaaS Project\trading-alerts-saas-public\davintrade-ui-design-stack\frontend-design\page-inventories.xlsx"
OUTPUT_BASE_DIR = r"D:\SaaS Project\trading-alerts-saas-public\davintrade-ui-design-stack\frontend-design\screenshot-image"
COUNTRY_CODES = ['gb', 'in', 'ng', 'pk', 'vn', 'id', 'th', 'za', 'tr', 'us', 'eu', 'jp']
CONCURRENCY = 6

def sanitize_filename(name):
    name = re.sub(r'[/\\?%*:|"<>]', '-', name)
    name = re.sub(r'-+', '-', name).strip('-')
    return name if name else 'home'

async def process_task(sem, context, task_info, total_count, progress_counter, lock):
    cc, folder_suffix, item_no, url, dest_file, filename = task_info
    async with sem:
        page = await context.new_page()
        success = False
        try:
            await page.goto(url, wait_until='networkidle', timeout=30000)
            await page.wait_for_timeout(600)
            await page.screenshot(path=dest_file, full_page=True)
            success = True
        except Exception as e:
            try:
                await page.goto(url, wait_until='load', timeout=30000)
                await page.wait_for_timeout(1000)
                await page.screenshot(path=dest_file, full_page=True)
                success = True
            except Exception as e2:
                async with lock:
                    print(f"FAILED [{cc.upper()}] Row {item_no:02d} ({url}): {e2}")
        finally:
            await page.close()

        async with lock:
            progress_counter[0] += 1
            curr = progress_counter[0]
            if success:
                print(f"[{curr}/{total_count}] Captured [{cc.upper()}] Row {item_no:02d} -> {cc}{folder_suffix}/{filename}")

        return success

async def main():
    folder_suffix = sys.argv[1] if len(sys.argv) > 1 else '-8'
    if not folder_suffix.startswith('-'):
        folder_suffix = '-' + folder_suffix

    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb.active

    for cc in COUNTRY_CODES:
        dir_path = os.path.join(OUTPUT_BASE_DIR, f"{cc}{folder_suffix}")
        os.makedirs(dir_path, exist_ok=True)

    col_to_cc = {}
    for c in range(2, 14):
        cc_val = str(sheet.cell(row=1, column=c).value or '').strip().lower()
        if cc_val in COUNTRY_CODES:
            col_to_cc[c] = cc_val

    tasks_info = []
    for r in range(2, sheet.max_row + 1):
        no_val = sheet.cell(row=r, column=1).value
        try:
            item_no = int(no_val)
        except (ValueError, TypeError):
            item_no = r - 1

        for c, cc in col_to_cc.items():
            url = sheet.cell(row=r, column=c).value
            if not url or not str(url).startswith('http'):
                continue
            url = str(url).strip()
            parsed = urllib.parse.urlparse(url)
            path_parts = [p for p in parsed.path.split('/') if p and p != cc]
            slug = '-'.join(path_parts) if path_parts else 'home'
            clean_slug = sanitize_filename(slug)

            filename = f"{item_no:02d}-{clean_slug}.png"
            dest_file = os.path.join(OUTPUT_BASE_DIR, f"{cc}{folder_suffix}", filename)
            tasks_info.append((cc, folder_suffix, item_no, url, dest_file, filename))

    total = len(tasks_info)
    print(f"Total screenshots queued ({folder_suffix}): {total}")
    start_time = time.time()

    sem = asyncio.Semaphore(CONCURRENCY)
    progress_counter = [0]
    lock = asyncio.Lock()

    async with async_playwright() as p:
        browser = await p.chromium.launch(channel='chrome', headless=True)
        context = await browser.new_context(
            viewport={'width': 1440, 'height': 900},
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, Gecko) Chrome/120.0.0.0 Safari/537.36'
        )

        futures = [process_task(sem, context, task, total, progress_counter, lock) for task in tasks_info]
        results = await asyncio.gather(*futures)

        await browser.close()

    successes = sum(1 for r in results if r)
    failures = total - successes
    elapsed = time.time() - start_time
    print(f"\n==========================================")
    print(f"FINISHED in {elapsed:.2f} seconds.")
    print(f"Total: {total} | Success: {successes} | Failed: {failures}")
    print(f"==========================================")

if __name__ == '__main__':
    asyncio.run(main())
